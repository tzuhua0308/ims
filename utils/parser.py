from __future__ import annotations
import re
from datetime import date, datetime, timedelta
from typing import Optional
import pandas as pd
import xlrd


# ── openpyxl 相容層（讓 .xlsx 也能用 xlrd 風格的 API）────────────────────────

class _XlsxSheet:
    def __init__(self, ws):
        self._data = [[cell.value for cell in row] for row in ws.iter_rows()]

    @property
    def nrows(self) -> int:
        return len(self._data)

    @property
    def ncols(self) -> int:
        return max((len(r) for r in self._data), default=0)

    def cell_value(self, r: int, c: int):
        try:
            return self._data[r][c]
        except IndexError:
            return None

    def row_values(self, r: int) -> list:
        try:
            return list(self._data[r])
        except IndexError:
            return []


class _XlsxBook:
    def __init__(self, wb):
        self._wb = wb

    def sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    def sheet_by_name(self, name: str) -> _XlsxSheet:
        return _XlsxSheet(self._wb[name])

PURCHASE_MAIN = "進項-可扣抵發票(進貨)"
SALES_SHEET   = "銷項"
EXPENSE_SHEET = "進項-發票(費用雜項)"
PRODUCT_SHEET = "國碼"


# ── 日期轉換 ────────────────────────────────────────────────────────────────

def to_date(val) -> Optional[date]:
    if val is None or (isinstance(val, float) and val == 0):
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, float) and val > 1000:
        try:
            return (date(1899, 12, 30) + timedelta(days=int(val)))
        except Exception:
            return None
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(val[:len(fmt.replace("%Y","0000").replace("%m","00").replace("%d","00"))], fmt).date()
            except Exception:
                pass
        # YYYY-MM-DD with time
        m = re.match(r"(\d{4}-\d{2}-\d{2})", val)
        if m:
            try:
                return datetime.strptime(m.group(1), "%Y-%m-%d").date()
            except Exception:
                pass
        # YYYYMM/D or YYYYMM/DD (e.g. 202511/3 → 2025-11-03)
        m2 = re.match(r"^(\d{4})(\d{2})/(\d{1,2})$", val)
        if m2:
            try:
                return date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
            except Exception:
                pass
    return None


def _safe_num(val) -> Optional[float]:
    try:
        v = float(val)
        return v if v == v else None  # NaN check
    except Exception:
        return None


# ── 年份 / 期別偵測 ──────────────────────────────────────────────────────────

def detect_year_period(filename: str) -> tuple[Optional[int], Optional[int]]:
    """從檔名 '115年01-02月...' 解析 (西元年, 期別)"""
    m = re.search(r"(\d+)年(\d+)-\d+月", filename)
    if m:
        roc = int(m.group(1))
        start_month = int(m.group(2))
        year = roc + 1911
        period = (start_month - 1) // 2 + 1
        return year, period
    return None, None


# ── 進貨解析 ─────────────────────────────────────────────────────────────────

def _sheet_declared_nums(ws) -> list[float]:
    """回傳工作表第 0 列所有 > 1000 的浮點數（可能的驗算值）。"""
    row0 = ws.row_values(0) if ws.nrows > 0 else []
    return [v for v in row0 if isinstance(v, float) and v > 1000]


def _closest(candidates: list[float], target: float) -> Optional[float]:
    """從候選清單中找最接近 target 的值。"""
    if not candidates:
        return None
    return min(candidates, key=lambda v: abs(v - target))


def _parse_purchase_sheet(ws, sheet_name: str, year: int, period: int) -> list[dict]:
    rows = []
    for r in range(1, ws.nrows):  # skip header row 0
        try:
            date_val  = ws.cell_value(r, 1)
            inv_type  = str(ws.cell_value(r, 2)).strip()
            inv_no    = str(ws.cell_value(r, 3)).strip()
            code      = str(ws.cell_value(r, 4)).strip()
            prod_name = str(ws.cell_value(r, 5)).strip()
            unit_price= _safe_num(ws.cell_value(r, 6))
            qty       = _safe_num(ws.cell_value(r, 7))
            vendor    = str(ws.cell_value(r, 8)).strip()
            vendor_tax= str(ws.cell_value(r, 9)).strip()
            amount    = _safe_num(ws.cell_value(r, 10))
        except IndexError:
            continue

        d = to_date(date_val)
        if d is None:
            continue
        if not amount or amount <= 0:
            continue

        rows.append({
            "year": year,
            "period": period,
            "date": d.isoformat(),
            "invoice_type": inv_type or None,
            "invoice_no": inv_no or None,
            "code": code or None,
            "product_name": prod_name or None,
            "unit_price": unit_price,
            "qty": qty,
            "vendor_name": vendor or None,
            "vendor_tax": vendor_tax or None,
            "amount": amount,
            "source_sheet": sheet_name,
        })
    return rows


def parse_purchases(wb: xlrd.Book, year: int, period: int) -> list[dict]:
    sheet_names = wb.sheet_names()
    if PURCHASE_MAIN not in sheet_names:
        return []
    ws = wb.sheet_by_name(PURCHASE_MAIN)
    return _parse_purchase_sheet(ws, PURCHASE_MAIN, year, period)


# ── 銷貨解析 ─────────────────────────────────────────────────────────────────

def parse_sales(wb: xlrd.Book, year: int, period: int) -> list[dict]:
    if SALES_SHEET not in wb.sheet_names():
        return []
    ws = wb.sheet_by_name(SALES_SHEET)
    rows = []
    for r in range(1, ws.nrows):
        try:
            machine   = str(ws.cell_value(r, 0)).strip()
            date_val  = ws.cell_value(r, 1)
            inv_no    = str(ws.cell_value(r, 2)).strip()
            code      = str(ws.cell_value(r, 3)).strip()
            prod_name = str(ws.cell_value(r, 4)).strip()
            qty       = _safe_num(ws.cell_value(r, 5))
            amount    = _safe_num(ws.cell_value(r, 6))
        except IndexError:
            continue

        if not amount or amount <= 0:
            continue
        d = to_date(date_val)

        rows.append({
            "year": year,
            "period": period,
            "machine_no": machine or None,
            "date": d.isoformat() if d else None,
            "invoice_no": inv_no or None,
            "code": code or None,
            "product_name": prod_name or None,
            "qty": qty,
            "untaxed_amount": amount,
        })
    return rows


# ── 費用解析 ─────────────────────────────────────────────────────────────────

def parse_expenses(wb: xlrd.Book, year: int, period: int) -> list[dict]:
    if EXPENSE_SHEET not in wb.sheet_names():
        return []
    ws = wb.sheet_by_name(EXPENSE_SHEET)
    rows = []
    for r in range(1, ws.nrows):
        try:
            date_val  = ws.cell_value(r, 1)
            inv_type  = str(ws.cell_value(r, 2)).strip()
            inv_no    = str(ws.cell_value(r, 3)).strip()
            content   = str(ws.cell_value(r, 4)).strip()
            vendor    = str(ws.cell_value(r, 5)).strip()
            vendor_tax= str(ws.cell_value(r, 6)).strip()
            # col 10 = 含稅額（實際支出金額）
            amount    = _safe_num(ws.cell_value(r, 10)) if ws.ncols > 10 else None
        except IndexError:
            continue

        if not amount or amount <= 0:
            continue
        d = to_date(date_val)

        rows.append({
            "year": year,
            "period": period,
            "date": d.isoformat() if d else None,
            "invoice_type": inv_type or None,
            "invoice_no": inv_no or None,
            "content": content or None,
            "vendor_name": vendor or None,
            "vendor_tax": vendor_tax or None,
            "amount": amount,
        })
    return rows


# ── 報廢解析（獨立 Excel）────────────────────────────────────────────────────

def parse_scraps_xlsx(file_bytes: bytes, year: int, period: int) -> list[dict]:
    """
    報廢 Excel 欄位：日期 / 代號 / 張數 / 原因 / 損失金額 / 備註
    """
    import io
    df = pd.read_excel(io.BytesIO(file_bytes), header=0)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {
        "日期": "date", "代號": "code", "張數": "qty",
        "原因": "reason", "損失金額": "loss", "備註": "note",
        "商品名稱": "product_name",
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

    rows = []
    for _, row in df.iterrows():
        loss = _safe_num(row.get("loss"))
        if not loss or loss <= 0:
            continue
        d = to_date(row.get("date"))
        rows.append({
            "year": year,
            "period": period,
            "date": d.isoformat() if d else None,
            "code": str(row.get("code", "") or "").strip() or None,
            "product_name": str(row.get("product_name", "") or "").strip() or None,
            "qty": _safe_num(row.get("qty")),
            "reason": str(row.get("reason", "") or "").strip() or None,
            "loss": loss,
            "note": str(row.get("note", "") or "").strip() or None,
        })
    return rows


# ── 品項主表解析 ──────────────────────────────────────────────────────────────

def parse_products(wb: xlrd.Book) -> list[dict]:
    if PRODUCT_SHEET not in wb.sheet_names():
        return []
    ws = wb.sheet_by_name(PRODUCT_SHEET)
    rows = []
    for r in range(ws.nrows):
        code = str(ws.cell_value(r, 0)).strip()
        name = str(ws.cell_value(r, 1)).strip()
        if not code or not name:
            continue
        rows.append({"code": code, "name": name, "ref_price": 0})
    return rows


# ── 主入口 ────────────────────────────────────────────────────────────────────

def parse_tax_xls(file_bytes: bytes, filename: str, year: int, period: int) -> dict:
    """
    回傳 {
        'purchases': [...], 'sales': [...], 'expenses': [...],
        'checksums': {sheet_name: {'declared': float, 'parsed': float}} — 各工作表驗算比對
    }
    """
    import io
    if filename.lower().endswith(".xlsx"):
        import openpyxl
        raw = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        wb = _XlsxBook(raw)
    else:
        wb = xlrd.open_workbook(file_contents=file_bytes)

    purchases = parse_purchases(wb, year, period)
    sales     = parse_sales(wb, year, period)
    expenses  = parse_expenses(wb, year, period)
    products  = parse_products(wb)

    # 建立驗算比對資料
    checksums: dict = {}
    sheet_names = wb.sheet_names()

    # 進貨主表
    if PURCHASE_MAIN in sheet_names:
        ws = wb.sheet_by_name(PURCHASE_MAIN)
        parsed_total = sum(r["amount"] for r in purchases)
        declared_nums = _sheet_declared_nums(ws)
        declared = _closest(declared_nums, parsed_total)
        checksums[PURCHASE_MAIN] = {"declared": declared, "parsed": parsed_total, "count": len(purchases)}

    # 費用工作表
    if EXPENSE_SHEET in sheet_names:
        ws = wb.sheet_by_name(EXPENSE_SHEET)
        exp_total = sum(r["amount"] for r in expenses)
        declared_nums = _sheet_declared_nums(ws)
        declared = _closest(declared_nums, exp_total)
        checksums[EXPENSE_SHEET] = {"declared": declared, "parsed": exp_total, "count": len(expenses)}

    return {
        "purchases": purchases,
        "sales":     sales,
        "expenses":  expenses,
        "products":  products,
        "checksums": checksums,
    }
